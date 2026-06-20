"""Central configuration helpers for the data cleaning pipeline.

This module defines project paths, report locations, runtime thresholds,
logging setup, and filesystem validation utilities used across the project.
Keeping this logic in one place makes the scripts easier to maintain and helps
prevent unsafe operations such as writing into the raw data directory.
"""

from __future__ import annotations

import logging
import re
import sys
from datetime import date, datetime
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Sequence

import pandas as pd
from openpyxl import load_workbook


LOGGER_NAME = "eda_cleaning_pipeline"
DEFAULT_LOG_LEVEL = logging.INFO
DEFAULT_CHUNK_SIZE = 10_000


@dataclass(frozen=True)
class MissingValueThresholds:
    """Business rules for missing value recommendations."""

    simple_imputation_max_pct: float = 5.0
    imputation_with_flag_max_pct: float = 40.0
    evaluate_utility_max_pct: float = 70.0


@dataclass(frozen=True)
class SimilarityThresholds:
    """RapidFuzz score thresholds for variant classification."""

    accepted_min: int = 95
    review_min: int = 80


@dataclass(frozen=True)
class AppConfig:
    """Resolved paths and thresholds for the project."""

    project_root: Path
    data_dir: Path
    raw_dir: Path
    processed_dir: Path
    reports_dir: Path
    raw_dataset_path: Path
    eda_report_path: Path
    missing_values_report_path: Path
    duplicates_report_path: Path
    organization_variants_report_path: Path
    cleaning_mapping_report_path: Path
    final_report_path: Path
    cleaned_dataset_path: Path
    chunk_size: int
    missing_value_thresholds: MissingValueThresholds
    similarity_thresholds: SimilarityThresholds


@dataclass(frozen=True)
class SheetProfile:
    """Metadata collected for a workbook sheet."""

    name: str
    row_count: int
    column_count: int
    non_empty_headers: int
    revision_number: int
    headers: tuple[str, ...]
    data_like_header_score: int
    is_data_candidate: bool


def get_project_root() -> Path:
    """Return the repository root based on the current module location."""

    return Path(__file__).resolve().parents[2]


def build_config(chunk_size: int = DEFAULT_CHUNK_SIZE) -> AppConfig:
    """Build the application configuration object.

    Args:
        chunk_size: Number of rows to process per chunk when chunked operations
            are supported.

    Returns:
        A fully resolved :class:`AppConfig` instance.

    Raises:
        ValueError: If the chunk size is not a positive integer.
    """

    if chunk_size <= 0:
        raise ValueError("chunk_size must be a positive integer")

    project_root = get_project_root()
    data_dir = project_root / "data"
    raw_dir = data_dir / "raw"
    processed_dir = data_dir / "processed"
    reports_dir = project_root / "reports"

    raw_dataset_path = (
        raw_dir
        / "2026-05-13_donnees-ouvertes_divulgation-octrois-subventions-et-contributions.xlsx"
    )

    return AppConfig(
        project_root=project_root,
        data_dir=data_dir,
        raw_dir=raw_dir,
        processed_dir=processed_dir,
        reports_dir=reports_dir,
        raw_dataset_path=raw_dataset_path,
        eda_report_path=reports_dir / "eda_report.md",
        missing_values_report_path=reports_dir / "missing_values_report.csv",
        duplicates_report_path=reports_dir / "duplicates_report.csv",
        organization_variants_report_path=reports_dir / "organization_variants_report.csv",
        cleaning_mapping_report_path=reports_dir / "cleaning_mapping.csv",
        final_report_path=reports_dir / "final_report.md",
        cleaned_dataset_path=processed_dir / "fichier_nettoye.csv",
        chunk_size=chunk_size,
        missing_value_thresholds=MissingValueThresholds(),
        similarity_thresholds=SimilarityThresholds(),
    )


def ensure_directories(paths: Iterable[Path]) -> None:
    """Create directories if they do not already exist.

    Args:
        paths: Directory paths to create.

    Raises:
        OSError: If a directory cannot be created.
    """

    for path in paths:
        path.mkdir(parents=True, exist_ok=True)


def initialize_project_structure(config: AppConfig) -> None:
    """Ensure required output directories exist.

    The raw directory is created for convenience, but this function never
    copies, edits, or moves the original dataset.
    """

    ensure_directories(
        [
            config.data_dir,
            config.raw_dir,
            config.processed_dir,
            config.reports_dir,
        ]
    )


def validate_input_dataset(config: AppConfig) -> None:
    """Ensure the expected source workbook exists and is readable.

    Args:
        config: Resolved application configuration.

    Raises:
        FileNotFoundError: If the workbook is missing.
        IsADirectoryError: If the configured workbook path points to a
            directory instead of a file.
    """

    dataset_path = config.raw_dataset_path

    if not dataset_path.exists():
        raise FileNotFoundError(
            "Source dataset not found. Expected workbook at "
            f"'{dataset_path}'. Place the original Excel file in data/raw/ "
            "without renaming it."
        )

    if dataset_path.is_dir():
        raise IsADirectoryError(
            f"Expected a workbook file but found a directory at '{dataset_path}'."
        )


def validate_output_path(output_path: Path, config: AppConfig) -> None:
    """Reject unsafe output targets.

    Args:
        output_path: Proposed output file path.
        config: Resolved application configuration.

    Raises:
        ValueError: If the path points inside the raw data directory or matches
            the original dataset path.
    """

    resolved_output = output_path.resolve()
    resolved_raw_dir = config.raw_dir.resolve()
    resolved_raw_dataset = config.raw_dataset_path.resolve()

    if resolved_output == resolved_raw_dataset:
        raise ValueError("Refusing to overwrite the original raw dataset.")

    if resolved_raw_dir in resolved_output.parents:
        raise ValueError(
            f"Unsafe output path '{output_path}': generated files must not be "
            "written inside data/raw/."
        )


def configure_logging(
    logger_name: str = LOGGER_NAME,
    level: int = DEFAULT_LOG_LEVEL,
) -> logging.Logger:
    """Create or reuse a project logger with a consistent console handler.

    Args:
        logger_name: Logger name to configure.
        level: Logging level, for example ``logging.INFO``.

    Returns:
        A configured :class:`logging.Logger`.
    """

    logger = logging.getLogger(logger_name)
    logger.setLevel(level)

    if not logger.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setLevel(level)
        handler.setFormatter(
            logging.Formatter(
                fmt="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
                datefmt="%Y-%m-%d %H:%M:%S",
            )
        )
        logger.addHandler(handler)

    logger.propagate = False
    return logger


def get_missing_value_recommendation(
    missing_pct: float,
    thresholds: MissingValueThresholds,
) -> str:
    """Map a missing-percentage value to the required recommendation label."""

    if missing_pct <= 0:
        return "Nothing"
    if missing_pct < thresholds.simple_imputation_max_pct:
        return "Simple Imputation"
    if missing_pct <= thresholds.imputation_with_flag_max_pct:
        return "Imputation + Missing Flag"
    if missing_pct <= thresholds.evaluate_utility_max_pct:
        return "Evaluate Utility"
    return "Possible Removal"


def classify_similarity_score(
    score: float,
    thresholds: SimilarityThresholds,
) -> str:
    """Classify a RapidFuzz score according to the project decision rules."""

    if score >= thresholds.accepted_min:
        return "accepted"
    if score >= thresholds.review_min:
        return "review"
    return "rejected"


def prepare_runtime(chunk_size: int = DEFAULT_CHUNK_SIZE) -> tuple[AppConfig, logging.Logger]:
    """Initialize directories, logging, and configuration for a script run.

    Args:
        chunk_size: Preferred processing chunk size.

    Returns:
        A tuple containing the resolved configuration and project logger.

    Raises:
        FileNotFoundError: If the source dataset is not available.
        OSError: If required directories cannot be created.
        ValueError: If the chunk size is invalid.
    """

    config = build_config(chunk_size=chunk_size)
    initialize_project_structure(config)
    validate_input_dataset(config)

    for output_path in (
        config.eda_report_path,
        config.missing_values_report_path,
        config.duplicates_report_path,
        config.organization_variants_report_path,
        config.cleaning_mapping_report_path,
        config.final_report_path,
        config.cleaned_dataset_path,
    ):
        validate_output_path(output_path, config)

    logger = configure_logging()
    logger.debug("Runtime prepared with project root: %s", config.project_root)
    return config, logger


def _coerce_header_name(value: object, position: int) -> str:
    """Convert a worksheet header cell into a safe column name."""

    if value is None:
        return f"unnamed_column_{position:02d}"

    header = str(value).strip()
    return header or f"unnamed_column_{position:02d}"


def inspect_workbook_sheets(config: AppConfig) -> list[SheetProfile]:
    """Inspect workbook sheets and collect metadata for sheet selection."""

    workbook = load_workbook(config.raw_dataset_path, read_only=True, data_only=True)
    profiles: list[SheetProfile] = []

    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            first_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                tuple(),
            )
            headers = tuple(
                _coerce_header_name(value, position)
                for position, value in enumerate(first_row, start=1)
            )
            header_tokens = [header.lower() for header in headers]
            non_empty_headers = sum(
                1 for value in first_row if value is not None and str(value).strip()
            )
            revision_match = re.search(r"rev-(\d+)", sheet_name.lower())
            revision_number = int(revision_match.group(1)) if revision_match else -1

            data_like_header_score = sum(
                1
                for header in header_tokens
                if any(
                    token in header
                    for token in (
                        "recipient",
                        "agreement",
                        "ref",
                        "owner_org",
                        "province",
                        "city",
                        "name",
                        "number",
                        "date",
                        "value",
                    )
                )
            )
            is_data_candidate = (
                worksheet.max_row > 1
                and non_empty_headers >= max(5, worksheet.max_column // 2)
                and data_like_header_score >= max(4, worksheet.max_column // 5)
            )

            profiles.append(
                SheetProfile(
                    name=sheet_name,
                    row_count=worksheet.max_row,
                    column_count=worksheet.max_column,
                    non_empty_headers=non_empty_headers,
                    revision_number=revision_number,
                    headers=headers,
                    data_like_header_score=data_like_header_score,
                    is_data_candidate=is_data_candidate,
                )
            )
    finally:
        workbook.close()

    return profiles


def select_primary_sheet(profiles: Sequence[SheetProfile]) -> SheetProfile:
    """Choose the most likely primary data sheet from workbook profiles."""

    if not profiles:
        raise ValueError("Workbook inspection returned no sheets.")

    return max(
        profiles,
        key=lambda profile: (
            int(profile.is_data_candidate),
            profile.data_like_header_score,
            profile.revision_number,
            profile.row_count,
            profile.column_count,
        ),
    )


def get_primary_sheet_profile(config: AppConfig) -> SheetProfile:
    """Inspect the workbook and return the best primary sheet candidate."""

    return select_primary_sheet(inspect_workbook_sheets(config))


def iter_sheet_chunks(
    config: AppConfig,
    sheet_name: str,
    chunk_size: int | None = None,
) -> Iterator[pd.DataFrame]:
    """Yield worksheet data as pandas DataFrames in memory-friendly chunks."""

    active_chunk_size = chunk_size or config.chunk_size
    if active_chunk_size <= 0:
        raise ValueError("chunk_size must be a positive integer")

    workbook = load_workbook(config.raw_dataset_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    row_iterator = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iterator)
        headers = [
            _coerce_header_name(value, position)
            for position, value in enumerate(header_row, start=1)
        ]
        chunk_rows: list[list[object]] = []

        for row in row_iterator:
            row_values = list(row[: len(headers)])
            if len(row_values) < len(headers):
                row_values.extend([None] * (len(headers) - len(row_values)))
            chunk_rows.append(row_values)

            if len(chunk_rows) >= active_chunk_size:
                yield pd.DataFrame(chunk_rows, columns=headers)
                chunk_rows = []

        if chunk_rows:
            yield pd.DataFrame(chunk_rows, columns=headers)
    finally:
        workbook.close()


def read_primary_sheet_sample(
    config: AppConfig,
    sample_size: int = 5_000,
) -> pd.DataFrame:
    """Read a bounded sample from the primary sheet for inference tasks."""

    if sample_size <= 0:
        raise ValueError("sample_size must be a positive integer")

    primary_sheet = get_primary_sheet_profile(config)
    collected_frames: list[pd.DataFrame] = []
    remaining = sample_size

    for chunk in iter_sheet_chunks(config, primary_sheet.name, chunk_size=min(sample_size, config.chunk_size)):
        if remaining <= 0:
            break

        collected_frames.append(chunk.head(remaining))
        remaining -= len(collected_frames[-1])

        if remaining <= 0:
            break

    if not collected_frames:
        return pd.DataFrame(columns=list(primary_sheet.headers))

    return pd.concat(collected_frames, ignore_index=True)


def is_text_like_value(value: object) -> bool:
    """Return whether a sampled cell value should be considered text-like."""

    return isinstance(value, str)


def is_numeric_like_value(value: object) -> bool:
    """Return whether a sampled cell value should be considered numeric-like."""

    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_date_like_value(value: object) -> bool:
    """Return whether a sampled cell value should be considered date-like."""

    return isinstance(value, (date, datetime))


def get_candidate_duplicate_columns(headers: Sequence[str]) -> list[str]:
    """Identify columns that are likely useful for duplicate detection."""

    priorities = (
        "ref_number",
        "agreement_number",
        "recipient_business_number",
        "recipient_legal_name",
        "recipient_legal_name_final",
        "recipient_operating_name",
        "agreement_title_en",
        "agreement_title_fr",
        "agreement_value",
        "agreement_start_date",
        "agreement_end_date",
        "owner_org",
    )
    available_headers = set(headers)
    return [column for column in priorities if column in available_headers]


def get_candidate_organization_columns(headers: Sequence[str]) -> list[str]:
    """Identify columns that may contain organization names."""

    priorities = (
        "recipient_legal_name",
        "recipient_legal_name_final",
        "recipient_legal_name_clean",
        "recipient_legal_name_clean_formula",
        "recipient_operating_name",
        "research_organization_name",
        "owner_org_title",
    )
    available_headers = set(headers)
    return [column for column in priorities if column in available_headers]


def select_primary_organization_column(headers: Sequence[str]) -> str | None:
    """Return the preferred organization name column for normalization work."""

    candidate_columns = get_candidate_organization_columns(headers)
    for preferred_column in (
        "recipient_legal_name",
        "recipient_legal_name_final",
        "recipient_legal_name_clean_formula",
        "recipient_legal_name_clean",
        "recipient_operating_name",
        "research_organization_name",
    ):
        if preferred_column in candidate_columns:
            return preferred_column
    return candidate_columns[0] if candidate_columns else None
